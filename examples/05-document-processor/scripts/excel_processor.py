#!/usr/bin/env python3
"""
Excel 文档处理脚本
支持读取工作表、获取单元格和分析数据
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("错误: 请安装 openpyxl 库")
    print("安装命令: pip install openpyxl")
    sys.exit(1)


def read_sheet(file_path, sheet_name=None):
    """读取工作表数据"""
    try:
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        return {
            "success": True,
            "sheet_name": ws.title,
            "data": data,
            "rows": len(data),
            "cols": len(data[0]) if data else 0
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def get_cell(file_path, cell_address, sheet_name=None):
    """获取指定单元格的值"""
    try:
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        cell_value = ws[cell_address].value
        
        return {
            "success": True,
            "cell": cell_address,
            "value": cell_value
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def analyze_data(file_path, sheet_name=None):
    """分析数据"""
    result = read_sheet(file_path, sheet_name)
    
    if not result["success"]:
        return result
    
    data = result["data"]
    
    # 简单的数据分析
    non_empty_rows = len([row for row in data if any(cell for cell in row)])
    
    return {
        "success": True,
        "sheet_name": result["sheet_name"],
        "total_rows": result["rows"],
        "total_cols": result["cols"],
        "non_empty_rows": non_empty_rows
    }


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("错误: 参数不足")
        print("用法: python excel_processor.py <action> <file_path> [options]")
        print("action: read_sheet, get_cell, analyze_data")
        sys.exit(1)
    
    action = sys.argv[1]
    file_path = sys.argv[2]
    
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)
    
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
    
    if action == "read_sheet":
        result = read_sheet(file_path, sheet_name)
        if result["success"]:
            print(f"工作表: {result['sheet_name']}")
            print(f"行数: {result['rows']}, 列数: {result['cols']}")
            print("\n数据:")
            for row in result["data"][:10]:  # 只显示前10行
                print(row)
        else:
            print(f"错误: {result['error']}")
            sys.exit(1)
    
    elif action == "get_cell":
        if len(sys.argv) < 4:
            print("错误: 请提供单元格地址")
            sys.exit(1)
        cell_address = sys.argv[3]
        result = get_cell(file_path, cell_address, sheet_name)
        if result["success"]:
            print(f"单元格 {result['cell']}: {result['value']}")
        else:
            print(f"错误: {result['error']}")
            sys.exit(1)
    
    elif action == "analyze_data":
        result = analyze_data(file_path, sheet_name)
        if result["success"]:
            print(f"工作表: {result['sheet_name']}")
            print(f"总行数: {result['total_rows']}")
            print(f"总列数: {result['total_cols']}")
            print(f"非空行数: {result['non_empty_rows']}")
        else:
            print(f"错误: {result['error']}")
            sys.exit(1)
    
    else:
        print(f"错误: 未知的操作类型: {action}")
        sys.exit(1)


